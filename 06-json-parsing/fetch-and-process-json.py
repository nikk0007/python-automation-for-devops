import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def fetch_json(url):
    try:
        # Send a GET request to the URL
        response = requests.get(url)
        
        # Raise an error if the response code is not 200 (OK)
        response.raise_for_status()

        # Parse the response as JSON
        data = response.json()

        # Return the JSON data
        return data

    except requests.exceptions.RequestException as e:
        # Handle any HTTP errors or exceptions
        print(f"An error occurred: {e}")
        return None

# URL of the JSON file
url = "http://3.92.57.155/employees.json"

# Fetch the JSON data
json_data = fetch_json(url)

if json_data:
    # Create a dictionary to hold department-wise data
    departments = {}

    # Loop through employees and organize by department
    for emp in json_data['employees']:
        dept = emp['department']
        # Combine address fields
        emp['office_address'] = ', '.join([emp['branch_office_address']['street'],
                                           emp['branch_office_address']['city'],
                                           emp['branch_office_address']['state'],
                                           emp['branch_office_address']['zip'],
                                           emp['branch_office_address']['country']])
        # Remove the original address field as it's no longer needed
        emp.pop('branch_office_address', None)

        # Prepare the row data for Excel
        row = {
            "Employee#": emp['employee_number'],
            "Name": emp['name'],
            "Designation": emp['designation'],
            "DOJ": emp['date_of_joining'],
            "Office Address": emp['office_address'],
            "Phone": emp['phone_number'],
            "Rating": emp['rating']
        }

        # Add employee data to the respective department
        if dept not in departments:
            departments[dept] = []
        departments[dept].append(row)

    # Create a Pandas Excel writer
    with pd.ExcelWriter('employee_details.xlsx', engine='openpyxl') as writer:
        # Loop through departments and create a sheet for each
        for dept, employees in departments.items():
            # Convert to DataFrame
            df = pd.DataFrame(employees)
            # Write to Excel sheet
            df.to_excel(writer, sheet_name=dept, index=False)

    # Load the workbook to apply conditional formatting
    wb = load_workbook('employee_details.xlsx')

    # Colors for conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Loop through departments and apply formatting and adjust column width
    for dept in departments.keys():
        sheet = wb[dept]
        
        # Loop through each row to apply conditional formatting for the 'Rating' column
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"G{row}"]  # Assuming 'Rating' is the 7th column
            if cell.value == 1:
                cell.fill = red_fill
            elif cell.value == 5:
                cell.fill = green_fill
        
        # Adjust the width of each column based on the length of the data in each column
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
            for cell in col:
                try:
                    # Calculate the maximum length of the cell values in this column
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Set the column width based on the maximum length of the content
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

    # Save the formatted workbook
    wb.save('employee_details.xlsx')

    print("Excel file created successfully.")
else:
    print("Failed to retrieve JSON data.")
